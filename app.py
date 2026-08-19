from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import os
import time
import re
import requests
from openpyxl import load_workbook
import shutil

app = Flask(__name__)
CORS(app, expose_headers=["Content-Disposition"])

# ================= CONFIGURAÇÕES =================
API_KEY = "a83c8826-5816-4975-afe3-f3597de54081-c770d61a-86d2-4862-a756-e1f04b30bf6c"
ARQUIVO_MODELO = "SUPERMERCADO PONTO CERTO_CADASTRO (1).xlsx"
# =================================================

os.makedirs("./temp", exist_ok=True)

def consultar_cnpj(cnpj):
    url = f"https://api.cnpja.com/office/{cnpj}?registrations=PR"
    headers = {"Authorization": API_KEY}
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json(), None
        # Códigos 401, 402, 403 e 429 geralmente indicam limite de requisições ou falta de pagamento/créditos
        elif response.status_code in [401, 402, 403, 429]:
            return None, "creditos"
        else:
            print(f"Erro na API. Código: {response.status_code}")
            return None, "servidor"
    except Exception as e:
        print(f"Erro na conexão com API: {e}")
        return None, "servidor"

def formatar_cnpj(cnpj_string):
    cnpj = re.sub(r'\D', '', str(cnpj_string)).zfill(14)
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def formatar_cep(cep_string):
    cep = re.sub(r'\D', '', str(cep_string)).zfill(8)
    return f"{cep[:2]}.{cep[2:5]}-{cep[5:]}"

def formatar_data(data_string):
    if not data_string: return ""
    partes = str(data_string).split('-')
    if len(partes) == 3: return f"{partes[2]}/{partes[1]}/{partes[0]}"
    return data_string

def gerar_excel(cnpj_limpo, pasta_destino, nome_rep="", codigo_rep=""):
    dados, erro_api = consultar_cnpj(cnpj_limpo)
    
    if erro_api: 
        return None, erro_api

    empresa = dados.get('company', {})
    razao_social = empresa.get('name', '')
    nome_fantasia = dados.get('alias') or razao_social
    data_fundacao = formatar_data(dados.get('founded', ''))

    inscricoes = dados.get('registrations', [])
    inscricao_estadual = ""
    for ie in inscricoes:
        if ie.get('state') == 'PR' and ie.get('enabled'):
            inscricao_estadual = ie.get('number', '')
            break

    endereco = dados.get('address', {})
    rua = endereco.get('street', '')
    numero = endereco.get('number', '')
    bairro = endereco.get('district', '')
    cep = formatar_cep(endereco.get('zip', ''))
    cidade = endereco.get('city', '')
    estado = endereco.get('state', '')
    endereco_completo = f"{rua}, {numero}"

    emails = dados.get('emails', [])
    email = emails[0].get('address', '') if emails else ''

    telefones = dados.get('phones', [])
    telefone = f"({telefones[0].get('area')}) {telefones[0].get('number')}" if telefones else ''

    membros = empresa.get('members', [])

    try:
        wb = load_workbook(ARQUIVO_MODELO)
        ws = wb.active

        celulas_limpar = ['H7', 'M8', 'H8', 'H9', 'M9', 'H10', 'M10', 'H11', 'M11', 'P11', 'H12', 'H13', 'M13', 'H15', 'M15']
        for celula in celulas_limpar:
            ws[celula] = None

        for row in range(20, 25):
            ws[f'C{row}'] = None
            ws[f'J{row}'] = None

        ws['H7'] = razao_social
        ws['M8'] = data_fundacao
        ws['H8'] = nome_fantasia
        ws['H9'] = formatar_cnpj(cnpj_limpo)
        ws['M9'] = inscricao_estadual
        ws['H10'] = endereco_completo
        ws['M10'] = cep
        ws['H11'] = bairro
        ws['M11'] = cidade
        ws['P11'] = estado
        ws['H12'] = telefone
        ws['H13'] = email
        ws['M13'] = email

        if nome_rep:
            ws['H15'] = nome_rep
        if codigo_rep:
            ws['M15'] = codigo_rep

        linha_socio = 20
        for socio in membros[:5]:
            pessoa = socio.get('person', {})
            ws[f'C{linha_socio}'] = pessoa.get('name', '')
            ws[f'J{linha_socio}'] = pessoa.get('taxId', '')
            linha_socio += 1

        nome_arquivo_seguro = re.sub(r'[\\/*?:"<>|]', "", str(nome_fantasia)).strip()
        if not nome_arquivo_seguro:
            nome_arquivo_seguro = f"Sem_Nome_{cnpj_limpo}"

        caminho_salvamento = os.path.join(pasta_destino, f"{nome_arquivo_seguro}_cadastro.xlsx")
        wb.save(caminho_salvamento)
        
        return caminho_salvamento, None
    except Exception as e:
        print(f"Erro ao manipular o Excel: {e}")
        return None, "servidor"

# ================= ROTAS DA API =================

@app.route('/gerar_unico', methods=['POST'])
def rota_unico():
    dados = request.json
    cnpj_limpo = dados.get('cnpj')
    nome_rep = dados.get('nome_representante', '')
    codigo_rep = dados.get('codigo_representante', '')
    
    caminho_arquivo, erro = gerar_excel(cnpj_limpo, "./temp", nome_rep, codigo_rep)
    
    if erro == "creditos":
        return jsonify({"erro": "creditos"}), 400
    elif erro == "servidor" or not caminho_arquivo:
        return jsonify({"erro": "servidor"}), 500
        
    if os.path.exists(caminho_arquivo):
        return send_file(caminho_arquivo, as_attachment=True, download_name=os.path.basename(caminho_arquivo))
    else:
        return jsonify({"erro": "servidor"}), 500


@app.route('/gerar_multiplos', methods=['POST'])
def rota_multiplos():
    dados = request.json
    lista_cnpjs = dados.get('cnpjs', [])
    nome_rep = dados.get('nome_representante', '')
    codigo_rep = dados.get('codigo_representante', '')
    
    pasta_lote = f"./temp/lote_{int(time.time())}"
    os.makedirs(pasta_lote, exist_ok=True)
    
    for cnpj in lista_cnpjs:
        caminho_arquivo, erro = gerar_excel(cnpj, pasta_lote, nome_rep, codigo_rep)
        
        if erro == "creditos":
            # Se bater o limite de créditos no meio do lote, cancela tudo e avisa
            return jsonify({"erro": "creditos"}), 400
        elif erro == "servidor":
            # Aqui você pode escolher se ele cancela tudo ou só pula o CNPJ problemático.
            # No momento, deixei pra cancelar pra evitar planilhas incompletas.
            return jsonify({"erro": "servidor"}), 500
            
        time.sleep(0.5) 
        
    caminho_zip_base = f"./temp/cadastros_lote"
    shutil.make_archive(caminho_zip_base, 'zip', pasta_lote)
    
    return send_file(f"{caminho_zip_base}.zip", as_attachment=True, download_name='cadastros_multiplos.zip')
